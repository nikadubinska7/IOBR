from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


RUN_ID = "9897"
MANIFEST_NAME = ".iobr_generated_manifest.json"


FILTER_COLUMNS = {
    "season": "CUSTOMER_CONFIRMED_SEASON_YEAR_CD",
    "sold_to": "SOLD_TO_CUSTOMER_NBR",
    "d_account": "D_ACCOUNT",
    "document_type": "SO_DOCUMENT_TYPE_GROUP_DESC",
    "distribution_method": "DISTRIBUTION_METHOD_CD",
}


ISSUE_QUANTITY_COMMENTS = {
    "Check rejection",
    "D-Group quantity is 0",
    "D-Group quantity is higher",
    "Nike quantity is higher",
}


OK_DATE_COMMENTS = {"OK", "On time or early", None, ""}


@dataclass(frozen=True)
class Table:
    name: str
    headers: list[str | None]
    rows: list[tuple]

    @property
    def index(self) -> dict[str, int]:
        return {header: idx for idx, header in enumerate(self.headers) if header is not None}


@dataclass(frozen=True)
class Batch:
    account: str
    sold_to_number: str
    sold_to_name: str


@dataclass(frozen=True)
class ReportDefinition:
    file_label: str
    source_sheet: str
    output_sheet: str
    columns: list[str] | None
    predicate: Callable[[tuple, dict[str, int]], bool]
    account_only: bool = False


def is_blank(value) -> bool:
    return value is None or value == ""


def numeric_zero(value) -> bool:
    return value in (None, "", 0)


def contains_cancel(value) -> bool:
    return "cancel" in str(value or "").casefold()


def clean_filename_part(value: str) -> str:
    value = str(value or "").strip()
    value = re.sub(r'[<>:"/\\\\|?*]+', "-", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" .") or "Unknown"


def row_value(row: tuple, idx: dict[str, int], column: str):
    pos = idx.get(column)
    if pos is None or pos >= len(row):
        return None
    return row[pos]


def read_workbook_tables(workbook_path: str | Path) -> dict[str, Table]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    tables: dict[str, Table] = {}
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(iterator))
        except StopIteration:
            headers = []
            rows = []
        else:
            rows = [tuple(row[: len(headers)]) for row in iterator]
        tables[worksheet.title] = Table(worksheet.title, headers, rows)
    workbook.close()
    return tables


def get_filter_metadata(workbook_path: str | Path) -> dict:
    tables = read_workbook_tables(workbook_path)
    options: dict[str, set[tuple[str, str]]] = {key: set() for key in FILTER_COLUMNS}
    sold_to_names: dict[str, str] = {}
    accounts_by_sold_to: dict[str, set[str]] = {}

    for table in tables.values():
        idx = table.index
        for row in table.rows:
            for key, column in FILTER_COLUMNS.items():
                if column not in idx:
                    continue
                value = row_value(row, idx, column)
                if is_blank(value):
                    continue
                text = str(value)
                if key == "sold_to":
                    sold_to_name = row_value(row, idx, "SOLD_TO_CUSTOMER_NM")
                    if not is_blank(sold_to_name):
                        sold_to_names[text] = str(sold_to_name)
                else:
                    options[key].add((text, text))

            sold_to = row_value(row, idx, "SOLD_TO_CUSTOMER_NBR")
            account = row_value(row, idx, "D_ACCOUNT")
            if not is_blank(sold_to) and not is_blank(account):
                accounts_by_sold_to.setdefault(str(sold_to), set()).add(str(account))

    for number, name in sold_to_names.items():
        options["sold_to"].add((number, f"{number} - {name}"))

    return {
        "options": {
            key: sorted(values, key=lambda item: item[1].casefold())
            for key, values in options.items()
        },
        "accounts_by_sold_to": {
            sold_to: sorted(accounts)
            for sold_to, accounts in accounts_by_sold_to.items()
        },
    }


def get_filter_options(workbook_path: str | Path) -> dict[str, list[tuple[str, str]]]:
    return get_filter_metadata(workbook_path)["options"]


def selected_set(selected_filters: dict[str, Iterable[str]], key: str) -> set[str]:
    return {str(value) for value in selected_filters.get(key, []) if str(value) != ""}


def matches_selected_value(value, selected: set[str]) -> bool:
    return not selected or str(value) in selected


def matches_global_filters(
    row: tuple,
    idx: dict[str, int],
    selected_filters: dict[str, Iterable[str]],
    *,
    account_only: bool = False,
) -> bool:
    keys = ("d_account",) if account_only else tuple(FILTER_COLUMNS)
    for key in keys:
        column = FILTER_COLUMNS[key]
        values = selected_set(selected_filters, key)
        if not values or column not in idx:
            continue
        if not matches_selected_value(row_value(row, idx, column), values):
            return False
    return True


def report_line_matched(row: tuple, idx: dict[str, int]) -> bool:
    return numeric_zero(row_value(row, idx, "REJECT_QTY"))


def report_country_of_origin(row: tuple, idx: dict[str, int]) -> bool:
    return row_value(row, idx, "COMMENT_COO") == "Check"


def report_style_id(row: tuple, idx: dict[str, int]) -> bool:
    return row_value(row, idx, "COMMENT_PROD_CD") == "Check" and numeric_zero(
        row_value(row, idx, "REJECT_QTY")
    )


def report_ihd_vs_crd(row: tuple, idx: dict[str, int]) -> bool:
    return (
        row_value(row, idx, "COMMENT_CRD_VS_IHD") not in OK_DATE_COMMENTS
        and numeric_zero(row_value(row, idx, "REJECT_QTY"))
    )


def report_cancellations(row: tuple, idx: dict[str, int]) -> bool:
    return (
        row_value(row, idx, "COMMENT_QUANTITY_PO_LEVEL") == "Check rejection"
        and row_value(row, idx, "REPORTING_CONFIRMED_QTY_PO_LEVEL") == 0
    )


def report_po_not_received(row: tuple, idx: dict[str, int]) -> bool:
    return row_value(row, idx, "COMMENTS") == "Check OFOA status"


def report_qty_by_size(row: tuple, idx: dict[str, int]) -> bool:
    return (
        row_value(row, idx, "COMMENT_QUANTITY_PO_LEVEL") in ISSUE_QUANTITY_COMMENTS
        and numeric_zero(row_value(row, idx, "REJECT_QTY"))
    )


def report_nike_data_only(row: tuple, idx: dict[str, int]) -> bool:
    return not contains_cancel(row_value(row, idx, "SO_STATUS_L2_DESC"))


REPORTS = [
    ReportDefinition(
        file_label="Line_Matched",
        source_sheet="Line_Matched",
        output_sheet="Line_Matched",
        columns=None,
        predicate=report_line_matched,
    ),
    ReportDefinition(
        file_label="Country-of-Origin",
        source_sheet="Line_Matched",
        output_sheet="COO",
        columns=[
            "D_ACCOUNT",
            "D_BUYER",
            "D_TO_BE_DELIVERED_AT",
            "D_ARTICLE_NO",
            "D_DZ",
            "SOLD_TO_CUSTOMER_NM",
            "CUSTOMER_REQUESTED_DELIVERY_DT",
            "CUSTOMER_CONFIRMED_SEASON_YEAR_CD",
            "PRODUCT_CD",
            "STYLE_NM",
            "DIVISION_DESC",
            "CATEGORY_DESC",
            "GENDER_DESC",
            "AGE_DESC",
            "COUNTRY_OF_MANUFACTURE_CODE",
            "D_COO",
            "PO7",
            "PO12",
            "SO_HEADER_NBR",
            "SO_ITEM_NBR",
            "SO_DOCUMENT_TYPE_GROUP_DESC",
            "DISTRIBUTION_METHOD_CD",
        ],
        predicate=report_country_of_origin,
    ),
    ReportDefinition(
        file_label="IHDvsCRD",
        source_sheet="Line_Matched",
        output_sheet="IHD vs CRD",
        columns=[
            "D_ORDER_DATE",
            "D_ACCOUNT",
            "D_BUYER",
            "D_TO_BE_DELIVERED_AT",
            "D_ARTICLE_NO",
            "D_DZ",
            "SOLD_TO_CUSTOMER_NM",
            "CUSTOMER_CONFIRMED_DELIVERY_DT",
            "CUSTOMER_REQUESTED_DELIVERY_DT",
            "D_IHD",
            "CUSTOMER_CONFIRMED_SEASON_YEAR_CD",
            "PRODUCT_CD",
            "STYLE_NM",
            "DIVISION_DESC",
            "CATEGORY_DESC",
            "GENDER_DESC",
            "AGE_DESC",
            "FIRST_PRODUCT_OFFER_DT",
            "LAUNCH_DT",
            "SO_STATUS_L1_DESC",
            "SO_STATUS_L2_DESC",
            "PO7",
            "PO12",
            "SO_HEADER_NBR",
            "SO_ITEM_NBR",
            "SO_DOCUMENT_TYPE_GROUP_DESC",
            "ORDER_TYPE",
            "DISTRIBUTION_METHOD_CD",
            "DEMAND_HEADER_CREATE_DT",
            "LAUNCH_CD_ITEM",
            "CRD_VS_IHD",
            "COMMENT_CRD_VS_IHD",
        ],
        predicate=report_ihd_vs_crd,
    ),
    ReportDefinition(
        file_label="Style-ID",
        source_sheet="Line_Matched",
        output_sheet="Product Code",
        columns=[
            "D_ORDER_DATE",
            "D_ACCOUNT",
            "D_BUYER",
            "D_TO_BE_DELIVERED_AT",
            "D_ARTICLE_NO",
            "D_DZ",
            "SOLD_TO_CUSTOMER_NM",
            "CUSTOMER_CONFIRMED_DELIVERY_DT",
            "CUSTOMER_REQUESTED_DELIVERY_DT",
            "CUSTOMER_CONFIRMED_SEASON_YEAR_CD",
            "PRODUCT_CD",
            "D_PROD CD",
            "STYLE_NM",
            "DIVISION_DESC",
            "CATEGORY_DESC",
            "GENDER_DESC",
            "AGE_DESC",
            "FIRST_PRODUCT_OFFER_DT",
            "LAUNCH_DT",
            "SO_STATUS_L1_DESC",
            "SO_STATUS_L2_DESC",
            "PO7",
            "PO12",
            "SO_HEADER_NBR",
            "SO_ITEM_NBR",
            "SO_DOCUMENT_TYPE_GROUP_DESC",
            "ORDER_TYPE",
            "DISTRIBUTION_METHOD_CD",
            "DEMAND_HEADER_CREATE_DT",
            "COMMENT_PROD_CD",
        ],
        predicate=report_style_id,
    ),
    ReportDefinition(
        file_label="Cancellations",
        source_sheet="Line_Matched",
        output_sheet="Cancellation (full item)",
        columns=[
            "D_ORDER_DATE",
            "D_ACCOUNT",
            "D_BUYER",
            "SO_ACTIVE_IND",
            "D_TO_BE_DELIVERED_AT",
            "D_ARTICLE_NO",
            "D_DZ",
            "SOLD_TO_CUSTOMER_NBR",
            "SOLD_TO_CUSTOMER_NM",
            "CUSTOMER_CONFIRMED_DELIVERY_DT",
            "CUSTOMER_REQUESTED_DELIVERY_DT",
            "D_IHD",
            "CUSTOMER_CONFIRMED_SEASON_YEAR_CD",
            "PRODUCT_CD",
            "D_PROD CD",
            "STYLE_NM",
            "DIVISION_DESC",
            "CATEGORY_DESC",
            "GENDER_DESC",
            "AGE_DESC",
            "FIRST_PRODUCT_OFFER_DT",
            "LAUNCH_DT",
            "SO_STATUS_L1_DESC",
            "SO_STATUS_L2_DESC",
            "COUNTRY_OF_MANUFACTURE_CODE",
            "D_COO",
            "CUSTOMER_PO_NBR",
            "PO7",
            "PO12",
            "PO_SPLIT",
            "SO_HEADER_NBR",
            "SO_ITEM_NBR",
            "SO_DOCUMENT_TYPE_GROUP_DESC",
            "ORDER_TYPE",
            "DISTRIBUTION_METHOD_CD",
            "DEMAND_HEADER_CREATE_DT",
            "LAUNCH_CD_ITEM",
            "DELIVERED_QTY",
            "REPORTING_CONFIRMED_QTY",
            "ORDER_ENTRY_QTY",
            "D_QUANTITY",
            "REPORTING_CONFIRMED_QTY_PO_LEVEL",
            "ORDER_ENTRY_QTY_PO_LEVEL",
            "D_QUANTITY_PO_LEVEL",
            "COMMENT_QUANTITY_PO_LEVEL",
            "REJECT_QTY",
            "SO_REJECT_REASON_CD",
        ],
        predicate=report_cancellations,
    ),
    ReportDefinition(
        file_label="PO not received - Cancellations",
        source_sheet="Line_DGroup_Data_Only",
        output_sheet="Only at Snipes",
        columns=[
            "D_ORDER_DATE",
            "D_ACCOUNT",
            "D_BUYER",
            "D_TO_BE_DELIVERED_AT",
            "D_ARTICLE_NO",
            "D_DZ",
            "D_CUST PO",
            "D_QUANTITY",
            "D_PROD CD",
            "D_IHD",
            "COMMENTS",
        ],
        predicate=report_po_not_received,
        account_only=True,
    ),
    ReportDefinition(
        file_label="Qty by Size",
        source_sheet="Size_Matched",
        output_sheet="Quantity & Cancellation (size)",
        columns=None,
        predicate=report_qty_by_size,
    ),
    ReportDefinition(
        file_label="Nike Data Only",
        source_sheet="Line_Nike_Data_Only",
        output_sheet="Line_Nike_Data_Only",
        columns=None,
        predicate=report_nike_data_only,
    ),
]


def output_columns(report: ReportDefinition, table: Table) -> list[str]:
    if report.columns is None:
        return [header for header in table.headers if header is not None]
    return report.columns


def project_row(row: tuple, idx: dict[str, int], columns: list[str]) -> list:
    return [row_value(row, idx, column) for column in columns]


def determine_batches(
    tables: dict[str, Table],
    selected_filters: dict[str, Iterable[str]],
) -> list[Batch]:
    table = tables["Line_Matched"]
    idx = table.index
    batches: dict[tuple[str, str], Batch] = {}

    for row in table.rows:
        if not matches_global_filters(row, idx, selected_filters):
            continue

        account = row_value(row, idx, "D_ACCOUNT")
        sold_to_number = row_value(row, idx, "SOLD_TO_CUSTOMER_NBR")
        sold_to_name = row_value(row, idx, "SOLD_TO_CUSTOMER_NM")
        if is_blank(account) or is_blank(sold_to_number):
            continue

        account_text = str(account)
        sold_to_number_text = str(sold_to_number)
        key = (account_text, sold_to_number_text)
        batches[key] = Batch(
            account=account_text,
            sold_to_number=sold_to_number_text,
            sold_to_name=str(sold_to_name or "Unknown Sold-To"),
        )

    return sorted(
        batches.values(),
        key=lambda batch: (
            batch.sold_to_name.casefold(),
            batch.sold_to_number,
            batch.account,
        ),
    )


def row_matches_batch(row: tuple, idx: dict[str, int], batch: Batch, report: ReportDefinition) -> bool:
    if "D_ACCOUNT" in idx and row_value(row, idx, "D_ACCOUNT") != batch.account:
        return False
    if not report.account_only and "SOLD_TO_CUSTOMER_NBR" in idx:
        if row_value(row, idx, "SOLD_TO_CUSTOMER_NBR") != batch.sold_to_number:
            return False
    return True


def select_report_rows(
    table: Table,
    report: ReportDefinition,
    batch: Batch,
    selected_filters: dict[str, Iterable[str]],
) -> list[list]:
    idx = table.index
    columns = output_columns(report, table)
    selected_rows = []
    for row in table.rows:
        if not row_matches_batch(row, idx, batch, report):
            continue
        if not matches_global_filters(
            row,
            idx,
            selected_filters,
            account_only=report.account_only,
        ):
            continue
        if not report.predicate(row, idx):
            continue
        selected_rows.append(project_row(row, idx, columns))
    return selected_rows


def write_report(path: Path, sheet_name: str, columns: list[str], rows: list[list]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.append(columns)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"

    for col_idx, column in enumerate(columns, start=1):
        max_len = len(str(column))
        for row_idx in range(2, min(worksheet.max_row, 100) + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def manifest_path(output_dir: str | Path) -> Path:
    return Path(output_dir) / MANIFEST_NAME


def clear_previous_run(output_dir: str | Path) -> None:
    path = manifest_path(output_dir)
    if not path.exists():
        return

    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return

    root = Path(output_dir).resolve()
    for file_name in manifest.get("files", []):
        file_path = (root / file_name).resolve()
        if root in file_path.parents and file_path.exists():
            file_path.unlink()

    for folder_name in sorted(manifest.get("folders", []), key=len, reverse=True):
        folder_path = (root / folder_name).resolve()
        if root in folder_path.parents and folder_path.exists():
            try:
                folder_path.rmdir()
            except OSError:
                pass

    path.unlink(missing_ok=True)


def write_manifest(output_dir: str | Path, files: list[Path], folders: list[Path]) -> None:
    root = Path(output_dir).resolve()
    manifest = {
        "run_id": RUN_ID,
        "files": [str(path.resolve().relative_to(root)) for path in files],
        "folders": [str(path.resolve().relative_to(root)) for path in folders],
    }
    manifest_path(output_dir).write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def generate_reports(
    workbook_path: str | Path,
    output_dir: str | Path,
    selected_filters: dict[str, Iterable[str]],
    progress: Callable[[str], None] | None = None,
) -> dict[str, int]:
    def emit(message: str) -> None:
        if progress:
            progress(message)

    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    emit("Reading workbook...")
    tables = read_workbook_tables(workbook_path)

    missing_sheets = [report.source_sheet for report in REPORTS if report.source_sheet not in tables]
    if missing_sheets:
        raise ValueError(f"Missing required sheets: {', '.join(sorted(set(missing_sheets)))}")

    batches = determine_batches(tables, selected_filters)
    if not batches:
        raise ValueError("No account / sold-to combinations matched the selected filters.")

    created_files: list[Path] = []
    created_folders: set[Path] = set()
    report_counts = {report.file_label: 0 for report in REPORTS}

    for batch in batches:
        folder = output_root / clean_filename_part(
            f"{batch.sold_to_number} - {batch.sold_to_name}"
        )
        created_folders.add(folder)
        emit(f"Creating files for {batch.sold_to_name} / {batch.account}...")

        for report in REPORTS:
            table = tables[report.source_sheet]
            columns = output_columns(report, table)
            rows = select_report_rows(table, report, batch, selected_filters)
            filename = f"iOBR - {RUN_ID} - {report.file_label} {batch.account}.xlsx"
            path = folder / filename
            write_report(path, report.output_sheet, columns, rows)
            created_files.append(path)
            report_counts[report.file_label] += len(rows)

    write_manifest(output_root, created_files, sorted(created_folders))
    emit(f"Done. Created {len(created_files)} files in {len(created_folders)} folders.")
    return report_counts
